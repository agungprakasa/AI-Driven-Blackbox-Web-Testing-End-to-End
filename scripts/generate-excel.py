#!/usr/bin/env python3
"""Generator laporan Excel untuk API Testing — dengan History Perbaikan.

Usage:
    python scripts/generate-excel.py
"""
from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "reports" / "excel"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
CATEGORY_FILL = PatternFill("solid", fgColor="D6E4F0")
CATEGORY_FONT = Font(bold=True, size=10, name="Calibri")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
OK_FONT = Font(color="006100", bold=True, size=10, name="Calibri")
NOK_FILL = PatternFill("solid", fgColor="FFC7CE")
NOK_FONT = Font(color="9C0006", bold=True, size=10, name="Calibri")
ONT_FILL = PatternFill("solid", fgColor="BDD7EE")
ONT_FONT = Font(color="1F4E78", bold=True, size=10, name="Calibri")
NORMAL_FONT = Font(size=10, name="Calibri")
BOLD_FONT = Font(bold=True, size=10, name="Calibri")
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
DONE_FONT = Font(color="006100", size=10, name="Calibri")
OPEN_FILL = PatternFill("solid", fgColor="FFC7CE")
OPEN_FONT = Font(color="9C0006", size=10, name="Calibri")


def baca(path: Path) -> str:
    return path.read_text(encoding="utf-8") if path.exists() else ""


def parse_tests(tc_teks: str) -> list[dict]:
    tests = []
    for line in tc_teks.splitlines():
        s = line.strip()
        if s.startswith("|") and re.match(r"^\|\s*TC-[A-Z0-9]+-?\d{3}\b", s):
            cols = [c.strip() for c in s.strip("|").split("|")]
            if len(cols) < 7:
                continue
            m = re.match(r"TC-([A-Z0-9]+)-?", cols[0])
            tests.append({
                "id": cols[0],
                "judul": cols[1],
                "endpoint": cols[2],
                "request": cols[3],
                "expected": cols[4],
                "priority": cols[5],
                "status": cols[6].upper(),
                "modul": m.group(1) if m else "UNKNOWN",
            })
    return tests


def get_category(tc_id: str) -> str:
    prefix_map = {
        "P": "Normal Case", "N": "Abnormal Case", "B": "Boundary Value",
        "S": "Security", "PF": "Performance", "DV": "Data Validation",
        "E2E": "Integration / E2E",
    }
    for prefix, cat in prefix_map.items():
        if re.match(rf"^TC-{prefix}\d", tc_id):
            return cat
    return "Lainnya"


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    tc_teks = baca(ROOT / "docs" / "test-case.md")
    defect_teks = baca(ROOT / "docs" / "defect-list.md")

    tests = parse_tests(tc_teks)
    if not tests:
        print("ERROR: Tidak ada test case ditemukan")
        return 1

    total = len(tests)
    n_pass = sum(1 for t in tests if t["status"] == "PASS")
    n_fail = sum(1 for t in tests if t["status"] == "FAIL")
    defect_count = len([l for l in defect_teks.splitlines()
                        if l.strip().startswith("|") and re.match(r"^\|\s*DEF-\d+", l.strip())])

    wb = Workbook()
    today = date.today().strftime("%d/%m/%Y")

    # =========================================================
    # SHEET 1: RINGKASAN (dengan formula otomatis)
    # =========================================================
    ws = wb.active
    ws.title = "Ringkasan"
    ws.sheet_properties.tabColor = "1F4E78"

    ht = "'Hasil Test'"
    hist = "'History Perbaikan'"
    sf = f"{ht}!F:F"  # Status column
    sp = f"{ht}!H:H"  # Status Perbaikan column

    ringkasan_data = [
        ("LAPORAN PENGUJIAN API", ""),
        ("", ""),
        ("Informasi Umum", ""),
        ("Tanggal Laporan", today),
        ("Endpoint", "POST /getfeeLnDiscountNew"),
        ("Base URL", "http://10.29.41.37:8280/test/1.0.0"),
        ("Penguji", "Agung Prakasa"),
        ("", ""),
        ("METRIK PENGUJIAN", "NILAI"),
        ("Total Test Case", f"=COUNTA({ht}!B:B)-1"),
        ("Status OK", f'=COUNTIF({sf},"OK")'),
        ("Status NOK", f'=COUNTIF({sf},"NOK")'),
        ("Status ON TEST", f'=COUNTIF({sf},"ON TEST")'),
        ("Pass Rate", "=IF(B11>0,B11/(B11+B12),\"-\")"),
        ("", ""),
        ("METRIK PERBAIKAN", "NILAI"),
        ("Total Perlu Perbaikan", f'=COUNTIF({sf},"NOK")'),
        ("Sudah Diperbaiki", f'=COUNTIF({sp},"DONE")'),
        ("Belum Diperbaiki", f'=COUNTIF({sp},"OPEN")'),
        ("Progress Perbaikan", "=IF(B18>0,B19/B18,\"-\")"),
        ("", ""),
        ("RIWAYAT PERUBAHAN", "Jumlah"),
        ("Total Perubahan Status", f"=COUNTA({hist}!A:A)-1"),
        ("", ""),
        ("DEFECT", "Jumlah"),
        ("Total Defect", defect_count),
    ]
    for r, (label, nilai) in enumerate(ringkasan_data, start=1):
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=nilai)
        a.border = BORDER
        b.border = BORDER
        if r == 1:
            a.font = Font(bold=True, size=14, name="Calibri")
        elif label in ("Informasi Umum", "METRIK PENGUJIAN", "METRIK PERBAIKAN", "RIWAYAT PERUBAHAN", "DEFECT"):
            a.font = HEADER_FONT
            b.font = HEADER_FONT
            a.fill = HEADER_FILL
            b.fill = HEADER_FILL
        elif label:
            a.font = BOLD_FONT
            b.font = NORMAL_FONT
            if label == "Pass Rate" or label == "Progress Perbaikan":
                b.number_format = '0.0%'
        a.alignment = Alignment(vertical="center")
        b.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # =========================================================
    # SHEET 2: HASIL TEST (dengan kolom progress)
    # =========================================================
    ws2 = wb.create_sheet("Hasil Test")
    ws2.sheet_properties.tabColor = "4472C4"

    headers = [
        "No.", "ID", "Butir Uji", "Uraian Kegiatan",
        "Hasil\nPengujian", "Status\nPengujian",
        "Keterangan", "Status\nPerbaikan",
        "Tgl Ditemukan", "Tgl Diperbaiki", "Oleh"
    ]
    col_widths = [6, 12, 22, 38, 22, 12, 28, 14, 14, 14, 14]

    # Write headers
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws2.row_dimensions[1].height = 40

    for c, w in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # Data validation for Status Pengujian (col 6)
    dv_status = DataValidation(
        type="list", formula1='"OK,NOK,ON TEST"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid", error="Pilih: OK, NOK, ON TEST"
    )
    ws2.add_data_validation(dv_status)

    # Data validation for Status Perbaikan (col 8)
    dv_perbaikan = DataValidation(
        type="list", formula1='"OPEN,DONE,WIP,N/A"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid", error="Pilih: OPEN, DONE, WIP, N/A"
    )
    ws2.add_data_validation(dv_perbaikan)

    # Group tests by category
    categories = []
    current_cat = None
    for t in tests:
        cat = get_category(t["id"])
        if cat != current_cat:
            categories.append({"name": cat, "tests": []})
            current_cat = cat
        categories[-1]["tests"].append(t)

    row_num = 2
    no_counter = 0

    for cat_info in categories:
        cat_name = cat_info["name"]
        cat_tests = cat_info["tests"]

        # Category row (merged)
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=11)
        cat_cell = ws2.cell(row=row_num, column=1, value=cat_name)
        cat_cell.font = CATEGORY_FONT
        cat_cell.fill = CATEGORY_FILL
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 12):
            ws2.cell(row=row_num, column=c).border = BORDER
            ws2.cell(row=row_num, column=c).fill = CATEGORY_FILL
        ws2.row_dimensions[row_num].height = 22
        row_num += 1

        for t in cat_tests:
            no_counter += 1
            tc_id = t["id"]
            status = t["status"]
            ok_nok = "OK" if status == "PASS" else "NOK"

            # Keterangan
            keterangan = ""
            if status == "FAIL":
                if "S" in tc_id[:4]:
                    keterangan = "P: Perlu input validation"
                elif "N" in tc_id[:4]:
                    keterangan = "P: Perlu validasi input"
                elif "B" in tc_id[:4]:
                    keterangan = "P: Perlu boundary check"
                else:
                    keterangan = "P: Perlu perbaikan"

            # Status Perbaikan
            status_perbaikan = "DONE" if status == "PASS" else "OPEN"

            # Tgl Ditemukan (always today — test execution date)
            tgl_ditemukan = today

            # Tgl Diperbaiki (only if DONE)
            tgl_diperbaiki = today if status == "PASS" else ""

            # Oleh
            oleh = "Agung Prakasa" if status == "PASS" else ""

            # Uraian Kegiatan
            deskripsi = f"Request: {t['request']}\nExpected: {t['expected']}"

            # Write row
            ws2.cell(row=row_num, column=1, value=no_counter).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(row=row_num, column=2, value=tc_id).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(row=row_num, column=3, value=t["judul"]).alignment = Alignment(vertical="center", wrap_text=True)
            ws2.cell(row=row_num, column=4, value=deskripsi).alignment = Alignment(vertical="center", wrap_text=True)
            ws2.cell(row=row_num, column=5, value=f"Status: {status}").alignment = Alignment(vertical="center", wrap_text=True)

            # Status Pengujian (dropdown)
            sc = ws2.cell(row=row_num, column=6, value=ok_nok)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            if ok_nok == "OK":
                sc.fill, sc.font = OK_FILL, OK_FONT
            else:
                sc.fill, sc.font = NOK_FILL, NOK_FONT
            dv_status.add(sc)

            ws2.cell(row=row_num, column=7, value=keterangan).alignment = Alignment(vertical="center", wrap_text=True)

            # Status Perbaikan (dropdown)
            sp_cell = ws2.cell(row=row_num, column=8, value=status_perbaikan)
            sp_cell.alignment = Alignment(horizontal="center", vertical="center")
            if status_perbaikan == "DONE":
                sp_cell.fill, sp_cell.font = DONE_FILL, DONE_FONT
            else:
                sp_cell.fill, sp_cell.font = OPEN_FILL, OPEN_FONT
            dv_perbaikan.add(sp_cell)

            ws2.cell(row=row_num, column=9, value=tgl_ditemukan).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(row=row_num, column=10, value=tgl_diperbaiki).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(row=row_num, column=11, value=oleh).alignment = Alignment(horizontal="center", vertical="center")

            # Apply font and border
            for c in range(1, 12):
                cell = ws2.cell(row=row_num, column=c)
                if c not in (6, 8):
                    cell.font = NORMAL_FONT
                cell.border = BORDER

            ws2.row_dimensions[row_num].height = 30
            row_num += 1

    ws2.freeze_panes = "A2"

    # =========================================================
    # SHEET 3: HISTORY PERBAIKAN (log perubahan status)
    # =========================================================
    ws3 = wb.create_sheet("History Perbaikan")
    ws3.sheet_properties.tabColor = "FFC000"

    h_headers = [
        "Tanggal", "Test Case ID", "Butir Uji", "Kategori",
        "Status Sebelum", "Status Sesudah", "Tindakan", "Oleh", "Keterangan"
    ]
    h_widths = [14, 14, 25, 18, 14, 14, 16, 16, 35]

    for c, h in enumerate(h_headers, start=1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="BF8F00")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws3.row_dimensions[1].height = 35

    for c, w in enumerate(h_widths, start=1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # Pre-fill initial history for all test cases
    hist_row = 2
    for t in tests:
        tc_id = t["id"]
        status = t["status"]
        ok_nok = "OK" if status == "PASS" else "NOK"
        kategori = get_category(tc_id)

        ws3.cell(row=hist_row, column=1, value=today).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=2, value=tc_id).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=3, value=t["judul"]).alignment = Alignment(vertical="center", wrap_text=True)
        ws3.cell(row=hist_row, column=4, value=kategori).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=5, value="-").alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=6, value=ok_nok).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=7, value="Initial Test").alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=8, value="Agung Prakasa").alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=hist_row, column=9, value="Pengujian awal").alignment = Alignment(vertical="center", wrap_text=True)

        # Color status cells
        s6 = ws3.cell(row=hist_row, column=6)
        if ok_nok == "OK":
            s6.fill, s6.font = OK_FILL, OK_FONT
        else:
            s6.fill, s6.font = NOK_FILL, NOK_FONT

        for c in range(1, 10):
            cell = ws3.cell(row=hist_row, column=c)
            if not cell.font.bold:
                cell.font = NORMAL_FONT
            cell.border = BORDER
        hist_row += 1

    ws3.freeze_panes = "A2"

    # =========================================================
    # SHEET 4: DAFTAR DEFECT
    # =========================================================
    ws4 = wb.create_sheet("Daftar Defect")
    ws4.sheet_properties.tabColor = "FF0000"

    d_headers = ["ID", "Judul", "Severity", "Status", "Keterangan"]
    d_widths = [12, 35, 12, 20, 40]

    for c, h in enumerate(d_headers, start=1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for c, w in enumerate(d_widths, start=1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    defect_lines = [l.strip() for l in defect_teks.splitlines()
                    if l.strip().startswith("|") and re.match(r"^\|\s*DEF-\d+", l.strip())]
    for ri, line in enumerate(defect_lines, start=2):
        cols = [c.strip() for c in line.strip("|").split("|")]
        for ci in range(min(5, len(cols))):
            cell = ws4.cell(row=ri, column=ci + 1, value=cols[ci])
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws4.freeze_panes = "A2"

    # =========================================================
    # SAVE
    # =========================================================
    # Get endpoint name from test case
    endpoint_name = "getfeeLnDiscountNew"
    out = OUT_DIR / f"Laporan-{endpoint_name}-{date.today().isoformat()}.xlsx"
    wb.save(out)
    print(f"Laporan Excel dibuat: {out.relative_to(ROOT)}")
    print(f"Total: {total} test cases | PASS: {n_pass} | FAIL: {n_fail}")
    print(f"Sheets: Ringkasan, Hasil Test, History Perbaikan, Daftar Defect")
    return 0


if __name__ == "__main__":
    sys.exit(main())
